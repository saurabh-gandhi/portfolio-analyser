"""
Build the output Excel workbook with all 6 analysis sheets.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Optional
from .utils import bar_chart, fmt_lakhs


# ── Colour palette ────────────────────────────────────────────────────────────
C = {
    'nav':     'FF1F3864',
    'white':   'FFFFFFFF',
    'black':   'FF000000',
    'alt':     'FFF2F7FF',
    'title':   'FF2E75B6',
    'green_h': 'FFE2EFDA',
    'green_t': 'FF375623',
    'amber':   'FFFFF2CC',
    'warn':    'FFFCE4D6',
}

AC_COLORS = {
    'Indian Equity':                     '4472C4',
    'Indian Private Equity (Unlisted)':  '2E75B6',
    'US/Intl Equity':                    '70AD47',
    'Commercial Real Estate':            '5B9BD5',
    'Indian Govt Bonds':                 'ED7D31',
    'Indian Corp Debt / CDs':            'FFC000',
    'Gold':                              'BF9000',
    'Cash & Equivalents':                '92D050',
    'Cash & Equivalents (Arbitrage)':    'A9D18E',
}

GRP_COLORS = {
    'Indian Equity': {'header': 'FF2E75B6', 'l1': 'FF4472C4', 'l2': 'FFAFD8F8', 'l3': 'FFD6E4F7'},
    'US Equity':     {'header': 'FF375623', 'l1': 'FF70AD47', 'l2': 'FFC6E0B4', 'l3': 'FFE2EFDA'},
    'Debt':          {'header': 'FF833C00', 'l1': 'FFED7D31', 'l2': 'FFFCE4D6', 'l3': 'FFFFF2CC'},
    'Gold':          {'header': 'FF7F5F00', 'l1': 'FFBF9000', 'l2': 'FFFFD966', 'l3': 'FFFFF2CC'},
    'Cash':          {'header': 'FF375623', 'l1': 'FF70AD47', 'l2': 'FFC6E0B4', 'l3': 'FFE2EFDA'},
    'RE':            {'header': 'FF1F5C8B', 'l1': 'FF5B9BD5', 'l2': 'FFB4D3EE', 'l3': 'FFD6E4F7'},
}

FONT_NAME = 'Arial'


def _fl(hex_color: str) -> PatternFill:
    return PatternFill('solid', start_color=hex_color, end_color=hex_color)

def _bd() -> Border:
    s = Side(style='thin', color='FFD0D0D0')
    return Border(left=s, right=s, top=s, bottom=s)

def _font(sz=9, bold=False, color=C['black'], name=FONT_NAME) -> Font:
    return Font(name=name, size=sz, bold=bold, color=color)

def _al(h='left', v='center', wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _hdr_row(ws, row_num: int, headers: list, widths: Optional[list] = None,
             bg: str = C['nav'], height: int = 28):
    ws.row_dimensions[row_num].height = height
    for i, h in enumerate(headers, 1):
        c = ws.cell(row_num, i, h)
        c.font  = _font(9, True, C['white'])
        c.fill  = _fl(bg)
        c.alignment = _al('center', wrap=True)
        c.border = _bd()
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

def _title_bar(ws, text: str, cols: int, bg: str = C['title'], height: int = 24, sz: int = 10):
    ws.merge_cells(f'A1:{get_column_letter(cols)}1')
    c = ws['A1']
    c.value = text
    c.font  = _font(sz, True, C['white'])
    c.fill  = _fl(bg)
    c.alignment = _al('left')
    ws.row_dimensions[1].height = height
    ws.sheet_view.showGridLines = False

def _data_cell(ws, r, col, val, fmt=None, aln='left', bg=C['white'],
               bold=False, ft=C['black'], sz=9, fn=FONT_NAME, wrap=False):
    c = ws.cell(r, col, None if val == '' else val)
    c.fill   = _fl(bg)
    c.border = _bd()
    c.font   = Font(name=fn, size=sz, bold=bold, color=ft)
    if fmt and val not in ('', None):
        c.number_format = fmt
    c.alignment = Alignment(horizontal=aln, vertical='center', wrap_text=wrap)


class ExcelWriter:
    def __init__(self, total_portfolio: float):
        self.wb = Workbook()
        self.total = total_portfolio
        # Remove default sheet
        self.wb.remove(self.wb.active)

    # ── Sheet builders ────────────────────────────────────────────────────────

    def add_true_allocation(self, rollup_df: pd.DataFrame, exposure_df: pd.DataFrame):
        ws = self.wb.create_sheet('🏦 True Allocation')
        ws.sheet_view.showGridLines = False
        ws.column_dimensions['A'].width = 2

        # Title
        ws.merge_cells('B2:M2')
        c = ws['B2']
        c.value = '🏛  TRUE PORTFOLIO ALLOCATION  |  Full look-through: EPF, PPF, NPS, LIC, REITs, MFs, ETFs'
        c.font  = _font(13, True, C['white']); c.fill = _fl(C['title']); c.alignment = _al('left')
        ws.row_dimensions[2].height = 34

        ws.merge_cells('B3:M3')
        c = ws['B3']
        c.value = (f'Total Portfolio: {fmt_lakhs(self.total)}   |   '
                   f'{len(rollup_df)} true asset classes after look-through all wrappers')
        c.font = _font(8, False, 'FF444444'); c.fill = _fl('FFEEF4FB'); c.alignment = _al('left')
        ws.row_dimensions[3].height = 14

        row = 5
        # Section header
        ws.merge_cells(f'B{row}:K{row}')
        c = ws.cell(row, 2, 'TRUE ASSET CLASS BREAKDOWN')
        c.font = _font(10, True); c.fill = _fl(C['nav']); c.alignment = _al('center')
        for col in range(3, 12): ws.cell(row, col).fill = _fl(C['nav'])
        row += 1

        hdrs   = ['#', 'True Asset Class', '₹ Exposure', '% of Portfolio', 'Visual (4%=1█)', 'Notes']
        widths = [4, 44, 18, 14, 24, 55]
        _hdr_row(ws, row, hdrs, widths=None, bg=C['nav'])
        for i, w in enumerate(widths, 2):
            ws.column_dimensions[get_column_letter(i)].width = w

        for rank, (_, r) in enumerate(rollup_df.iterrows(), 1):
            row += 1
            ac   = str(r['True Asset Class'])
            rs   = float(r['Rs Exposure'])
            pct  = float(r['Pct of Total'])
            clr  = AC_COLORS.get(ac, 'AAAAAA')
            is_lic = 'lic' in ac.lower() or 'insurance' in ac.lower()
            bg   = C['amber'] if is_lic else (C['alt'] if rank % 2 == 0 else C['white'])
            bar  = bar_chart(pct, scale=4, width=25)

            cells = [
                (2,  rank,  '0',       'center', False, C['black'], FONT_NAME, 9),
                (3,  ac,    None,      'left',   True,  'FF'+clr,   FONT_NAME, 9),
                (4,  rs,    '₹#,##0',  'right',  False, C['black'], FONT_NAME, 9),
                (5,  pct/100, '0.00%', 'right',  False, C['black'], FONT_NAME, 9),
                (6,  bar,   None,      'left',   False, 'FF'+clr,   'Consolas', 8),
                (7,  '',    None,      'left',   False, C['black'], FONT_NAME, 9),
            ]
            for col, val, fmt, aln, bold, ft, fn, sz in cells:
                _data_cell(ws, row, col, val, fmt, aln, bg, bold, ft, sz, fn)

        # Total row
        row += 1
        for col, (val, fmt, aln) in enumerate([
            ('TOTAL PORTFOLIO', None, 'left'),
            (self.total, '₹#,##0', 'right'),
            (1.0, '0.00%', 'right'),
            ('', None, 'left'), ('', None, 'left'),
        ], 2):
            c = ws.cell(row, col, None if val == '' else val)
            c.fill = _fl(C['nav']); c.border = _bd()
            c.font = _font(9, True, C['white'])
            if fmt and val not in ('', None): c.number_format = fmt
            c.alignment = _al(aln)

    def add_sub_class_breakdown(self, sub_rollup_df: pd.DataFrame):
        """Add 🗂 Sub-Class Breakdown sheet."""
        ws = self.wb.create_sheet('🗂 Sub-Class Breakdown')
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A4'

        _title_bar(ws, '🗂  ASSET CLASS SUB-BREAKDOWN  |  Large/Mid/Small  •  Debt Duration  •  Gold Form  •  Cash Type',
                   7, height=28, sz=11)

        ws.merge_cells('A2:G2')
        c = ws['A2']
        c.value = f'Total Portfolio: {fmt_lakhs(self.total)}   |   All values weighted by actual ₹ allocation'
        c.font = _font(8, False, 'FF444444'); c.fill = _fl('FFEEF4FB'); c.alignment = _al('left')
        ws.row_dimensions[2].height = 14

        for i, w in enumerate([3, 44, 18, 14, 18, 22, 58], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Column headers
        _hdr_row(ws, 3, ['', 'Sub-Class', '₹ Exposure', '% of Asset Class',
                          '% of Total Portfolio', 'Bar (2%=1█)', 'Instruments / Notes'],
                 bg=C['nav'], height=26)

        def _section(start_row, group_label, sub_rows, grp_key):
            clrs = GRP_COLORS.get(grp_key, GRP_COLORS['Indian Equity'])
            r    = start_row
            grp_rs = sum(x[1] for x in sub_rows)
            grp_pct = grp_rs / self.total * 100

            # Header bar
            for col in range(1, 8): ws.cell(r, col).fill = _fl(clrs['header']); ws.cell(r, col).border = _bd()
            ws.merge_cells(f'B{r}:G{r}')
            c = ws.cell(r, 2, f'  {group_label}   —   {fmt_lakhs(grp_rs)}   ({grp_pct:.1f}% of portfolio)')
            c.font = _font(10, True, C['white']); c.fill = _fl(clrs['header']); c.alignment = _al('left')
            ws.row_dimensions[r].height = 22
            r += 1

            for idx, (sub, rs, pct_ac, pct_tot, note) in enumerate(sub_rows):
                bg = clrs['l2'] if idx % 2 == 0 else clrs['l3']
                ws.cell(r, 1).fill = _fl(bg); ws.cell(r, 1).border = _bd()
                for col, val, fmt, aln, bold, ft, fn, sz in [
                    (2, sub,           None,      'left',  True,  C['black'],           FONT_NAME,   9),
                    (3, rs,            '₹#,##0',  'right', False, C['black'],           FONT_NAME,   9),
                    (4, pct_ac/100,    '0.0%',    'right', False, C['black'],           FONT_NAME,   9),
                    (5, pct_tot/100,   '0.00%',   'right', False, C['black'],           FONT_NAME,   9),
                    (6, bar_chart(pct_tot,2), None,'left', False, 'FF'+clrs['l1'][2:], 'Consolas',  8),
                    (7, note,          None,      'left',  False, 'FF555555',           FONT_NAME,   8),
                ]:
                    _data_cell(ws, r, col, val, fmt, aln, bg, bold, ft, sz, fn)
                ws.row_dimensions[r].height = 15
                r += 1

            # Subtotal
            for col in range(1, 8): ws.cell(r, col).fill = _fl(clrs['l1']); ws.cell(r, col).border = _bd()
            for col, val, fmt, aln in [
                (2, f'  ↳ {group_label} TOTAL', None, 'left'),
                (3, grp_rs, '₹#,##0', 'right'),
                (4, 1.0,    '0.0%',   'right'),
                (5, grp_pct/100, '0.00%', 'right'),
            ]:
                c = ws.cell(r, col, val); c.fill = _fl(clrs['l1']); c.border = _bd()
                c.font = _font(9, True, C['white'])
                if fmt: c.number_format = fmt
                c.alignment = _al(aln)
            ws.row_dimensions[r].height = 18
            return r + 2

        # Group sub_rollup into buckets
        def _get(sub_class: str) -> float:
            mask = sub_rollup_df['Sub Class'] == sub_class
            if mask.any():
                return float(sub_rollup_df.loc[mask, 'Rs Exposure'].values[0])
            return 0.0

        def _get_like(pattern: str) -> float:
            mask = sub_rollup_df['Sub Class'].str.contains(pattern, case=False, na=False)
            return float(sub_rollup_df.loc[mask, 'Rs Exposure'].sum())

        row = 4

        # Indian Equity
        ie_lc = _get('Indian Equity — Large Cap')
        ie_mc = _get('Indian Equity — Mid Cap')
        ie_sc = _get('Indian Equity — Small Cap')
        ie_pe = _get('Indian Equity — Unlisted / Private Equity')
        ie_tot = ie_lc + ie_mc + ie_sc + ie_pe or 1
        row = _section(row, '🇮🇳 INDIAN EQUITY', [
            ('Large Cap (Nifty 50/100, Next 50)', ie_lc, ie_lc/ie_tot*100, ie_lc/self.total*100,
             'Index funds, NIFTYBEES, JUNIORBEES, NPS/EPF equity slice'),
            ('Mid Cap (Nifty Midcap 150)',         ie_mc, ie_mc/ie_tot*100, ie_mc/self.total*100,
             'MIDCAPETF, Zerodha LM250 (60%), Mirae L&M, mid-cap fund sleeves'),
            ('Small Cap',                           ie_sc, ie_sc/ie_tot*100, ie_sc/self.total*100,
             'Small cap funds, small-cap fund sleeves'),
            ('Unlisted / Private Equity',           ie_pe, ie_pe/ie_tot*100, ie_pe/self.total*100,
             'ESOPs — unlisted equity, valued at last known round'),
        ], 'Indian Equity')

        # US Equity
        us_tot = _get_like('US / Intl Equity')
        if us_tot > 0:
            row = _section(row, '🌎 US / INTL EQUITY', [
                ('Large / Mega Cap (Nasdaq 100 focus)', us_tot, 100.0, us_tot/self.total*100,
                 'Nasdaq 100 FOFs/ETFs, QQQ, NVDA, PPFCF overseas holdings'),
            ], 'US Equity')

        # Debt
        dlt_g  = _get_like('Debt — Govt Bonds, Long Term')
        dlt_l  = _get_like('Debt — LIC')
        dlt_c  = _get_like('Debt — Corp/Other, Long Term')
        dst_c  = _get_like('Debt — Corp/Other, Short Term')
        d_tot  = dlt_g + dlt_l + dlt_c + dst_c or 1
        debt_rows = []
        if dlt_g > 0: debt_rows.append(('Long Term — Govt Bonds (>5 yrs)', dlt_g, dlt_g/d_tot*100, dlt_g/self.total*100, 'PPF, EPF G-Secs/SDLs, NPS G-Secs'))
        if dlt_l > 0: debt_rows.append(('Long Term — LIC / Insurance', dlt_l, dlt_l/d_tot*100, dlt_l/self.total*100, 'LIC endowment/term policies — illiquid debt wrapper'))
        if dlt_c > 0: debt_rows.append(('Long Term — Corp / Other (3-10 yrs)', dlt_c, dlt_c/d_tot*100, dlt_c/self.total*100, 'EPF/NPS corporate bonds, REIT embedded debt'))
        if dst_c > 0: debt_rows.append(('Short Term — CDs / CPs / Bonds (<3 yrs)', dst_c, dst_c/d_tot*100, dst_c/self.total*100, 'Fund cash sleeves, short-dated bonds'))
        if debt_rows:
            row = _section(row, '🏛  DEBT', debt_rows, 'Debt')

        # Gold
        g_sgb = _get_like('Gold — Sovereign Gold Bond')
        g_phy = _get_like('Gold — Physical')
        g_etf = _get_like('Gold — ETF')
        g_tot = g_sgb + g_phy + g_etf or 1
        gold_rows = []
        if g_sgb > 0: gold_rows.append(('Sovereign Gold Bonds (Paper, Govt-backed, 2.5% interest)', g_sgb, g_sgb/g_tot*100, g_sgb/self.total*100, 'SGBs — earns 2.5% p.a. + gold appreciation'))
        if g_phy > 0: gold_rows.append(('Physical Gold (Jewellery / Coins / Bars)', g_phy, g_phy/g_tot*100, g_phy/self.total*100, 'No yield, making charges deducted on sale'))
        if g_etf > 0: gold_rows.append(('Gold ETF (Paper, exchange-traded, most liquid)', g_etf, g_etf/g_tot*100, g_etf/self.total*100, 'GOLDBEES etc. — can sell any day on exchange'))
        if gold_rows:
            row = _section(row, '🥇 GOLD', gold_rows, 'Gold')

        # Cash
        c_arb = _get_like('Arbitrage')
        c_bnk = _get_like('Cash — Bank')
        c_buf = _get_like('Cash & Equivalents — Fund Buffer')
        c_rnt = _get_like('Cash — Rental')
        c_tot = c_arb + c_bnk + c_buf + c_rnt or 1
        cash_rows = []
        if c_arb > 0: cash_rows.append(('Arbitrage Fund (Hedged, ~7-8% p.a., T+1)', c_arb, c_arb/c_tot*100, c_arb/self.total*100, 'Net equity direction ≈ 0. Earns cash-futures spread.'))
        if c_bnk > 0: cash_rows.append(('Bank / Savings Account', c_bnk, c_bnk/c_tot*100, c_bnk/self.total*100, 'Working capital / emergency fund'))
        if c_buf > 0: cash_rows.append(('Fund Cash Buffers (TREPS)', c_buf, c_buf/c_tot*100, c_buf/self.total*100, 'Settlement buffers in MFs/ETFs'))
        if c_rnt > 0: cash_rows.append(('Rental Deposit / Receivable', c_rnt, c_rnt/c_tot*100, c_rnt/self.total*100, ''))
        if cash_rows:
            row = _section(row, '💵 CASH & LIQUID', cash_rows, 'Cash')

        # Real Estate
        re_tot = _get_like('Commercial Real Estate')
        if re_tot > 0:
            row = _section(row, '🏢 COMMERCIAL REAL ESTATE (REITs)', [
                ('Grade A Office / Commercial (REIT equity component)', re_tot, 100.0, re_tot/self.total*100,
                 'REIT equity NAV portion (~80% of total REIT holding)'),
            ], 'RE')

        # Footer
        ws.merge_cells(f'A{row}:G{row}')
        c = ws.cell(row, 1, f'  PORTFOLIO TOTAL   {fmt_lakhs(self.total)}')
        c.font = _font(10, True, C['white']); c.fill = _fl(C['nav'])
        c.alignment = _al('left'); ws.row_dimensions[row].height = 22
        for col in range(2, 8): ws.cell(row, col).fill = _fl(C['nav']); ws.cell(row, col).border = _bd()

    def add_look_through_detail(self, exposure_df: pd.DataFrame):
        ws = self.wb.create_sheet('🔬 Look-Through Detail')
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A3'
        _title_bar(ws, f'LOOK-THROUGH DETAIL  |  Every asset mapped to true economic exposure  |  ₹{self.total:,.0f}', 7)

        hdrs   = ['Asset', 'Present Value (₹)', 'True Asset Class', 'Alloc %', '₹ Exposure', '% of Total', 'Notes']
        widths = [44, 18, 40, 10, 16, 14, 55]
        _hdr_row(ws, 2, hdrs, widths)

        for idx, (_, r) in enumerate(exposure_df.sort_values('Rs Exposure', ascending=False).iterrows(), 1):
            rn  = idx + 2
            ac  = str(r['True Asset Class'])
            clr = AC_COLORS.get(ac.split('(')[0].strip(), 'AAAAAA')
            is_lic  = 'lic' in ac.lower() or 'insurance' in ac.lower()
            is_ppaf = 'arbitrage' in str(r.get('Asset', '')).lower()
            bg = C['amber'] if is_lic else (C['warn'] if is_ppaf else (C['alt'] if idx % 2 == 0 else C['white']))

            for col, val, fmt, aln, bold, ft in [
                (1, r['Asset'],          None,      'left',  False, C['black']),
                (2, r['Present Value'],  '₹#,##0',  'right', False, C['black']),
                (3, ac,                  None,      'left',  True,  'FF'+clr),
                (4, r['Allocation Pct']/100, '0%',  'right', False, C['black']),
                (5, r['Rs Exposure'],    '₹#,##0',  'right', False, C['black']),
                (6, r['Rs Exposure']/self.total, '0.00%', 'right', False, C['black']),
                (7, '',                  None,      'left',  False, 'FF555555'),
            ]:
                _data_cell(ws, rn, col, val, fmt, aln, bg, bold, ft)
            ws.row_dimensions[rn].height = 14

    def add_stock_rollup(self, stock_df: pd.DataFrame, equity_pool: float):
        ws = self.wb.create_sheet('🔍 Stock Roll-Up')
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A3'
        n = len(stock_df)
        _title_bar(ws, f'STOCK-LEVEL BOTTOMS-UP  |  {n} unique stocks/assets  |  Equity Pool: {fmt_lakhs(equity_pool)}', 10)

        hdrs   = ['Rank', 'Stock / Asset', 'Sector', 'Instruments', 'Held In',
                  'Total ₹ Exposure', '% of Total Portfolio', '% of Equity Pool', 'Max %', 'Source']
        widths = [5, 40, 24, 12, 90, 18, 17, 15, 12, 26]
        _hdr_row(ws, 2, hdrs, widths, height=32)
        ws.row_dimensions[2].height = 32

        for idx, (_, r) in enumerate(stock_df.iterrows(), 1):
            rn = idx + 2
            bg = C['alt'] if idx % 2 == 0 else C['white']
            rs = float(r.get('Total ₹ Exposure', 0))
            for col, val, fmt, aln, bold in [
                (1,  int(r['Rank']),                         '0',      'center', False),
                (2,  r.get('Stock Name', ''),                None,     'left',   idx <= 10),
                (3,  r.get('Sector', ''),                    None,     'left',   False),
                (4,  int(r.get('Instruments', 0)),           '0',      'center', False),
                (5,  r.get('Held In', ''),                   None,     'left',   False),
                (6,  rs,                                     '₹#,##0', 'right',  False),
                (7,  r.get('Pct of Total Portfolio', 0)/100, '0.00%',  'right',  False),
                (8,  r.get('Pct of Equity Pool', 0)/100,     '0.00%',  'right',  False),
                (9,  r.get('Max % in Any Instrument', 0)/100,'0.00%',  'right',  False),
                (10, r.get('Source', ''),                    None,     'left',   False),
            ]:
                _data_cell(ws, rn, col, val, fmt, aln, bg, bold)

    def add_sector_breakdown(self, sector_df: pd.DataFrame):
        ws = self.wb.create_sheet('📊 Equity Sector Breakdown')
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A3'
        _title_bar(ws, 'EQUITY SECTOR BREAKDOWN  |  From bottoms-up stock-level analysis', 6)

        hdrs   = ['#', 'Sector', 'No. of Stocks', '₹ Exposure', '% of Equity Pool', '% of Total Portfolio']
        widths = [5, 34, 14, 20, 20, 20]
        _hdr_row(ws, 2, hdrs, widths)

        for rank, (_, r) in enumerate(sector_df.iterrows(), 1):
            rn = rank + 2
            bg = C['alt'] if rank % 2 == 0 else C['white']
            for col, val, fmt, aln, bold in [
                (1, rank,                                     '0',      'center', False),
                (2, r.get('Industry', r.get('Sector', '')),   None,     'left',   rank <= 3),
                (3, int(r.get('Stock_Count', r.get('No. of Stocks', 0))), '0', 'center', False),
                (4, float(r.get('Total_Rs', r.get('₹ Exposure', 0))), '₹#,##0', 'right', False),
                (5, float(r.get('Pct of Equity Pool', 0))/100, '0.00%', 'right', False),
                (6, float(r.get('Pct of Total Portfolio', 0))/100, '0.00%', 'right', False),
            ]:
                _data_cell(ws, rn, col, val, fmt, aln, bg, bold)

    def add_cash_holdings(self, cash_df: pd.DataFrame, total_instruments: float):
        ws = self.wb.create_sheet('💵 Cash & Debt Holdings')
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A4'
        _title_bar(ws, f'CASH & DEBT HOLDINGS  |  Per fund breakdown  |  ₹{cash_df["Weighted ₹ Exposure"].sum():,.0f} total', 7)

        ws.merge_cells('A2:G2')
        c = ws['A2']
        c.value = '⚠️  Arbitrage fund cash = futures margin/float — NOT freely deployable equity cash'
        c.font = _font(8, True, 'FF7F7F00'); c.fill = _fl(C['amber']); c.alignment = _al('left')
        ws.row_dimensions[2].height = 20

        for i, w in enumerate([44, 60, 26, 12, 20, 20, 16], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        hdrs = ['Instrument', 'Item / Security', 'Category', '% to NAV', 'Weighted ₹ Exposure', 'Instrument Value (₹)', 'Data As Of']
        _hdr_row(ws, 3, hdrs, height=26)

        for idx, (_, r) in enumerate(cash_df.iterrows(), 1):
            rn  = idx + 3
            is_ppaf = 'arbitrage' in str(r.get('Instrument', '')).lower()
            bg  = C['warn'] if is_ppaf else (C['alt'] if idx % 2 == 0 else C['white'])
            for col, val, fmt, aln in [
                (1, r.get('Instrument', ''),            None,      'left'),
                (2, r.get('Item', ''),                  None,      'left'),
                (3, r.get('Category', ''),              None,      'left'),
                (4, r.get('% to NAV', 0)/100,           '0.00%',   'right'),
                (5, r.get('Weighted ₹ Exposure', 0),    '₹#,##0',  'right'),
                (6, r.get('Instrument Value (₹)', 0),   '₹#,##0',  'right'),
                (7, r.get('Data As Of', ''),             None,      'center'),
            ]:
                _data_cell(ws, rn, col, val, fmt, aln, bg)

    def save(self, output_path: str):
        self.wb.save(output_path)
