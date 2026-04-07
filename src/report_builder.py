"""
report_builder.py
Builds the multi-sheet Excel analysis report.
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


FONT_NAME = "Arial"

# ── Colour palette ────────────────────────────────────────────────────────────
C = {
    "nav":      "FF1F3864",
    "white":    "FFFFFFFF",
    "black":    "FF000000",
    "alt":      "FFF2F7FF",
    "title":    "FF2E75B6",
    "green_h":  "FFE2EFDA",
    "green_ft": "FF375623",
    "amber":    "FFFFF2CC",
    "warn":     "FFFCE4D6",
    "eq_hdr":   "FF2E75B6",
    "us_hdr":   "FF375623",
    "debt_hdr": "FF833C00",
    "gold_hdr": "FF7F5F00",
    "cash_hdr": "FF375623",
    "re_hdr":   "FF1F5C8B",
}

ASSET_COLORS = {
    "Indian Equity":                             "4472C4",
    "Indian Equity — Large Cap":                 "4472C4",
    "Indian Equity — Mid Cap":                   "2E75B6",
    "Indian Equity — Small Cap":                 "1F4E79",
    "Indian Equity — Unlisted / Private":        "17375E",
    "US/Intl Equity":                            "70AD47",
    "US/Intl Equity — Large/Mega Cap":           "70AD47",
    "Commercial Real Estate":                    "5B9BD5",
    "Indian Govt Bonds":                         "ED7D31",
    "Indian Govt Bonds — Long Term":             "ED7D31",
    "Indian Corp Debt":                          "FFC000",
    "Gold":                                      "BF9000",
    "Gold — Sovereign Gold Bonds (Paper)":       "BF9000",
    "Gold — Physical":                           "7F5F00",
    "Gold — ETF (Paper)":                        "E2A000",
    "Cash & Equivalents (Arbitrage)":            "A9D18E",
    "Cash & Equivalents":                        "92D050",
    "Cash — Bank / Savings Account":             "70AD47",
    "Insurance / LIC Wrapper":                   "9E480E",
    "Debt — Insurance/LIC":                      "C55A11",
}


def _fl(c: str) -> PatternFill:
    return PatternFill("solid", start_color=c, end_color=c)

def _bd() -> Border:
    s = Side(style="thin", color="FFD0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def _al(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _hf(sz=9, bold=True, c=C["white"]) -> Font:
    return Font(name=FONT_NAME, size=sz, bold=bold, color=c)

def _cf(sz=9, bold=False, c=C["black"]) -> Font:
    return Font(name=FONT_NAME, size=sz, bold=bold, color=c)


def _write_hdr(ws, row: int, headers: list, widths: list = None,
               bg: str = C["nav"], height: int = 28):
    ws.row_dimensions[row].height = height
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h)
        c.font = _hf(); c.fill = _fl(bg); c.alignment = _al("center"); c.border = _bd()
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def _title_row(ws, text: str, cols: int, bg: str = C["title"], height: int = 22):
    ws.merge_cells(f"A1:{get_column_letter(cols)}1")
    c = ws["A1"]
    c.value = text
    c.font = Font(name=FONT_NAME, size=10, bold=True, color=C["white"])
    c.fill = _fl(bg); c.alignment = _al("left")
    ws.row_dimensions[1].height = height
    ws.sheet_view.showGridLines = False


def _dc(ws, r: int, col: int, val, fmt=None, aln="left",
        bg=C["white"], bold=False, ft=C["black"], sz=9, fn=FONT_NAME, wrap=False):
    """Write a single data cell."""
    c = ws.cell(r, col, None if val == "" else val)
    c.fill = _fl(bg); c.border = _bd()
    c.font = Font(name=fn, size=sz, bold=bold, color=ft)
    if fmt and val not in ("", None): c.number_format = fmt
    c.alignment = Alignment(horizontal=aln, vertical="center", wrap_text=wrap)


def _bar(pct: float, scale: float = 2) -> str:
    n = max(0, min(20, round(pct / scale)))
    return "█" * n + "░" * (20 - n)


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet builders
# ═══════════════════════════════════════════════════════════════════════════════

def _build_true_allocation(wb: Workbook, rollup_df: pd.DataFrame,
                            subclass_df: pd.DataFrame, total: float):
    ws = wb.create_sheet("🏦 True Allocation")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    # Banner
    ws.row_dimensions[2].height = 34
    ws.merge_cells("B2:O2")
    c = ws["B2"]
    c.value = "🏛  TRUE PORTFOLIO ALLOCATION  |  Full Look-Through of All Wrappers"
    c.font = Font(name=FONT_NAME, size=14, bold=True, color=C["white"])
    c.fill = _fl(C["title"]); c.alignment = _al("left")
    ws.merge_cells("B3:O3")
    c = ws["B3"]
    c.value = f"Total Portfolio: ₹{total/100000:.2f}L   |   All wrappers looked through: EPF, PPF, NPS, LIC, REITs, MFs, ETFs"
    c.font = Font(name=FONT_NAME, size=8, color="FF444444")
    c.fill = _fl("FFEEF4FB"); c.alignment = _al("left")
    ws.row_dimensions[3].height = 16

    row = 5
    ws.merge_cells(f"B{row}:K{row}")
    ws.cell(row, 2, "TRUE ASSET CLASS BREAKDOWN").font = _hf(10)
    ws.cell(row, 2).fill = _fl(C["nav"]); ws.cell(row, 2).alignment = _al("center")
    for col in range(3, 12): ws.cell(row, col).fill = _fl(C["nav"])

    row += 1
    _write_hdr(ws, row,
               ["#", "True Asset Class", "₹ Exposure", "% of Portfolio", "Visual (4%=1█)", "Notes"],
               bg=C["nav"])
    for i, w in enumerate([4, 42, 18, 14, 24, 55], 2):
        ws.column_dimensions[get_column_letter(i)].width = w

    for rank, (_, r) in enumerate(rollup_df.iterrows(), 1):
        row += 1
        ac = r.get("True Asset Class", "")
        rs = r.get("Rs Exposure", 0)
        pct = r.get("Pct of Total", 0)
        clr = ASSET_COLORS.get(ac, "AAAAAA")
        bg = C["amber"] if "LIC" in ac or "Insurance" in ac else (C["alt"] if rank % 2 == 0 else C["white"])
        bar_str = _bar(pct, 4)

        _dc(ws, row, 2, rank, "0", "center", bg)
        _dc(ws, row, 3, ac, None, "left", bg, True, "FF" + clr)
        _dc(ws, row, 4, rs, "₹#,##0", "right", bg)
        _dc(ws, row, 5, pct / 100, "0.00%", "right", bg)
        _dc(ws, row, 6, bar_str, None, "left", bg, False, "FF" + clr, 8, "Consolas")
        _dc(ws, row, 7, r.get("Notes", ""), None, "left", bg, False, "FF555555", 8, FONT_NAME, True)

    # Totals
    row += 1
    for j, (v, fmt, aln) in enumerate([
        ("TOTAL PORTFOLIO", None, "left"), (total, "₹#,##0", "right"),
        (1.0, "0.00%", "right"), ("", None, "left"), ("", None, "left")], 2):
        c = ws.cell(row, j, None if v == "" else v)
        c.fill = _fl(C["nav"]); c.border = _bd()
        c.font = Font(name=FONT_NAME, size=9, bold=True, color=C["white"])
        if fmt and v != "": c.number_format = fmt
        c.alignment = _al(aln)

    # Summary line
    row += 2
    eq_rs = rollup_df[rollup_df["True Asset Class"].str.contains("Equity", na=False)]["Rs Exposure"].sum()
    dbt_rs = rollup_df[rollup_df["True Asset Class"].str.contains("Debt|Bonds|Corp|Insurance|LIC", na=False)]["Rs Exposure"].sum()
    gld_rs = rollup_df[rollup_df["True Asset Class"].str.contains("Gold", na=False)]["Rs Exposure"].sum()
    csh_rs = rollup_df[rollup_df["True Asset Class"].str.contains("Cash", na=False)]["Rs Exposure"].sum()
    summary = (f"Equity: {eq_rs/total*100:.1f}%   |   Debt (incl. LIC): {dbt_rs/total*100:.1f}%   |   "
               f"Gold: {gld_rs/total*100:.1f}%   |   Cash+Arb: {csh_rs/total*100:.1f}%")
    ws.merge_cells(f"B{row}:K{row}")
    c = ws.cell(row, 2, summary)
    c.fill = _fl(C["green_h"]); c.border = _bd()
    c.font = Font(name=FONT_NAME, size=9, bold=True, color=C["green_ft"])
    c.alignment = _al("left"); ws.row_dimensions[row].height = 20
    for col in range(3, 12): ws.cell(row, col).fill = _fl(C["green_h"]); ws.cell(row, col).border = _bd()


def _build_subclass_sheet(wb: Workbook, subclass_df: pd.DataFrame, total: float):
    ws = wb.create_sheet("🗂 Sub-Class Breakdown")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "🗂  ASSET CLASS SUB-BREAKDOWN  |  Large/Mid/Small  •  Debt Duration  •  Gold Form  •  Cash Type"
    c.font = Font(name=FONT_NAME, size=12, bold=True, color=C["white"])
    c.fill = _fl(C["title"]); c.alignment = _al("left"); ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = f"Total Portfolio: ₹{total/100000:.2f}L   |   Weighted by actual ₹ allocation"
    c.font = Font(name=FONT_NAME, size=8, color="FF444444")
    c.fill = _fl("FFEEF4FB"); c.alignment = _al("left"); ws.row_dimensions[2].height = 14

    for i, w in enumerate([3, 42, 18, 14, 18, 22, 55], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _write_hdr(ws, 3, ["", "Sub-Class", "₹ Exposure", "% of Asset Class", "% of Total Portfolio",
                        "Bar (2%=1█)", "Instruments / Sources"], bg=C["nav"], height=26)

    # Group definitions
    groups = [
        ("🇮🇳 INDIAN EQUITY",    ["Indian Equity"],         C["eq_hdr"],   "2E75B6"),
        ("🌎 US / INTL EQUITY",   ["US/Intl Equity"],        C["us_hdr"],   "70AD47"),
        ("🏛  DEBT",              ["Debt", "Bonds", "Corp",
                                   "Indian Govt", "Insurance", "LIC"], C["debt_hdr"], "ED7D31"),
        ("🥇 GOLD",               ["Gold"],                  C["gold_hdr"], "BF9000"),
        ("💵 CASH & LIQUID",      ["Cash"],                  C["cash_hdr"], "92D050"),
        ("🏢 REAL ESTATE",        ["Real Estate", "Commercial"], C["re_hdr"], "5B9BD5"),
    ]

    row = 4
    for group_label, keywords, hdr_color, bar_color in groups:
        # Filter rows for this group
        mask = subclass_df["Sub Class"].apply(
            lambda x: any(kw.upper() in str(x).upper() for kw in keywords)
        )
        group_df = subclass_df[mask]
        if group_df.empty:
            continue

        group_rs = group_df["Rs Exposure"].sum()
        group_pct = group_rs / total * 100

        # Section header
        for col in range(1, 8): ws.cell(row, col).fill = _fl(hdr_color); ws.cell(row, col).border = _bd()
        ws.merge_cells(f"B{row}:G{row}")
        c = ws.cell(row, 2, f"  {group_label}   —   ₹{group_rs/100000:.2f}L   ({group_pct:.1f}% of total)")
        c.font = Font(name=FONT_NAME, size=10, bold=True, color=C["white"])
        c.fill = _fl(hdr_color); c.alignment = _al("left")
        ws.row_dimensions[row].height = 22; row += 1

        for idx, (_, r) in enumerate(group_df.iterrows()):
            sub = r["Sub Class"]
            rs = r["Rs Exposure"]
            pct_tot = r["Pct of Total"]
            pct_ac = rs / group_rs * 100 if group_rs > 0 else 0
            bg = "FFCFE2F3" if idx % 2 == 0 else "FFE8F4FD"

            ws.cell(row, 1).fill = _fl(bg); ws.cell(row, 1).border = _bd()
            _dc(ws, row, 2, sub,         None,      "left",  bg, True,  C["black"], 9)
            _dc(ws, row, 3, rs,          "₹#,##0",  "right", bg, False, C["black"], 9)
            _dc(ws, row, 4, pct_ac/100,  "0.0%",    "right", bg, False, C["black"], 9)
            _dc(ws, row, 5, pct_tot/100, "0.00%",   "right", bg, False, C["black"], 9)
            _dc(ws, row, 6, _bar(pct_tot, 2), None, "left",  bg, False, "FF"+bar_color, 8, "Consolas")
            _dc(ws, row, 7, "",          None,      "left",  bg, False, "FF555555", 8)
            ws.row_dimensions[row].height = 15; row += 1

        # Subtotal
        for col in range(1, 8): ws.cell(row, col).fill = _fl("FF" + bar_color); ws.cell(row, col).border = _bd()
        _dc(ws, row, 2, f"  ↳ {group_label} TOTAL", None, "left",  "FF"+bar_color, True,  C["white"], 9)
        _dc(ws, row, 3, group_rs, "₹#,##0", "right", "FF"+bar_color, True, C["white"], 9)
        _dc(ws, row, 5, group_pct/100, "0.00%", "right", "FF"+bar_color, True, C["white"], 9)
        ws.row_dimensions[row].height = 18; row += 2

    # Footer
    for col in range(1, 8): ws.cell(row, col).fill = _fl(C["nav"]); ws.cell(row, col).border = _bd()
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row, 1, f"  PORTFOLIO TOTAL   ₹{total/100000:.2f}L")
    c.font = Font(name=FONT_NAME, size=10, bold=True, color=C["white"])
    c.fill = _fl(C["nav"]); c.alignment = _al("left"); ws.row_dimensions[row].height = 22


def _build_stock_rollup(wb: Workbook, stock_df: pd.DataFrame, total: float):
    ws = wb.create_sheet("🔍 Stock Roll-Up")
    ws.freeze_panes = "A3"; ws.sheet_view.showGridLines = False
    eq_pool = stock_df["Total_Rs"].sum() if not stock_df.empty else 0
    _title_row(ws, f"STOCK-LEVEL BOTTOMS-UP  |  {len(stock_df)} unique stocks  |  "
               f"Equity Pool: ₹{eq_pool/100000:.1f}L  |  Total Portfolio: ₹{total/100000:.1f}L", 10)

    hdrs = ["Rank", "Stock / Asset", "Sector", "Instruments", "Held In",
            "Total ₹ Exposure", "% of Total Portfolio", "% of Equity Pool", "Max % in Any", "Source"]
    wids = [5, 40, 24, 12, 85, 18, 17, 16, 14, 26]
    _write_hdr(ws, 2, hdrs, bg=C["nav"], height=36)
    for i, w in enumerate(wids, 1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 36

    for idx, (_, r) in enumerate(stock_df.iterrows(), 1):
        rn = idx + 2; bg = C["alt"] if idx % 2 == 0 else C["white"]
        rs = r.get("Total_Rs", 0)
        pct_eq = r.get("Pct of Equity Pool", 0)
        pct_tot = r.get("Pct of Total Portfolio", 0)
        _dc(ws, rn, 1,  int(r.get("Rank", idx)), "0",      "center", bg, False)
        _dc(ws, rn, 2,  r.get("Norm", ""),        None,     "left",   bg, idx <= 10)
        _dc(ws, rn, 3,  r.get("Sector", ""),      None,     "left",   bg)
        _dc(ws, rn, 4,  int(r.get("Instruments_Count", 0)), "0", "center", bg)
        _dc(ws, rn, 5,  r.get("Held_In", ""),     None,     "left",   bg)
        _dc(ws, rn, 6,  rs,         "₹#,##0",     "right",  bg)
        _dc(ws, rn, 7,  pct_tot/100,"0.00%",       "right",  bg)
        _dc(ws, rn, 8,  pct_eq/100, "0.00%",       "right",  bg)
        _dc(ws, rn, 9,  r.get("Max_Pct", 0)/100, "0.00%",   "right",  bg)
        _dc(ws, rn, 10, r.get("Source_Types", ""), None,    "left",   bg)


def _build_sector_sheet(wb: Workbook, sector_df: pd.DataFrame, total: float):
    ws = wb.create_sheet("📊 Equity Sector Breakdown")
    ws.freeze_panes = "A3"; ws.sheet_view.showGridLines = False
    _title_row(ws, "EQUITY SECTOR BREAKDOWN  |  From bottoms-up stock analysis", 7)

    hdrs = ["#", "Sector", "No. of Stocks", "₹ Exposure", "% of Equity Pool", "% of Total Portfolio"]
    wids = [5, 34, 14, 20, 20, 20]
    _write_hdr(ws, 2, hdrs, bg=C["nav"], height=28)
    for i, w in enumerate(wids, 1): ws.column_dimensions[get_column_letter(i)].width = w

    for rank, (_, r) in enumerate(sector_df.iterrows(), 1):
        rn = rank + 2; bg = C["alt"] if rank % 2 == 0 else C["white"]
        _dc(ws, rn, 1, rank, "0", "center", bg)
        _dc(ws, rn, 2, r.get("Sector", ""), None, "left", bg, rank <= 3)
        _dc(ws, rn, 3, int(r.get("Stock_Count", 0)), "0", "center", bg)
        _dc(ws, rn, 4, r.get("Total_Rs", 0), "₹#,##0", "right", bg)
        _dc(ws, rn, 5, r.get("Pct of Equity Pool", 0)/100, "0.00%", "right", bg)
        _dc(ws, rn, 6, r.get("Pct of Total Portfolio", 0)/100, "0.00%", "right", bg)


def _build_loothrough_sheet(wb: Workbook, exposure_rows: list, total: float):
    ws = wb.create_sheet("🔬 Look-Through Detail")
    ws.freeze_panes = "A3"; ws.sheet_view.showGridLines = False
    _title_row(ws, f"LOOK-THROUGH DETAIL  |  Every asset mapped to true economic exposure  |  ₹{total:,.0f}", 8)

    hdrs = ["Asset / Instrument", "Present Value (₹)", "True Asset Class",
            "Allocation %", "₹ Exposure", "% of Total Portfolio", "Notes"]
    wids = [44, 18, 38, 10, 16, 14, 55]
    _write_hdr(ws, 2, hdrs, bg=C["nav"], height=28)
    for i, w in enumerate(wids, 1): ws.column_dimensions[get_column_letter(i)].width = w

    for idx, row in enumerate(sorted(exposure_rows, key=lambda r: r.rs_exposure, reverse=True), 1):
        rn = idx + 2
        ac = row.true_asset_class
        clr = ASSET_COLORS.get(ac, ASSET_COLORS.get(ac.split("—")[0].strip(), "AAAAAA"))
        is_lic = "LIC" in ac or "Insurance" in ac
        is_ppaf = "Arbitrage" in row.asset_name
        bg = C["amber"] if is_lic else (C["warn"] if is_ppaf else (C["alt"] if idx % 2 == 0 else C["white"]))

        _dc(ws, rn, 1, row.asset_name,         None,     "left",  bg)
        _dc(ws, rn, 2, row.present_value,       "₹#,##0", "right", bg)
        _dc(ws, rn, 3, ac,                      None,     "left",  bg, True, "FF"+clr)
        _dc(ws, rn, 4, row.allocation_pct/100,  "0%",     "right", bg)
        _dc(ws, rn, 5, row.rs_exposure,         "₹#,##0", "right", bg)
        _dc(ws, rn, 6, row.rs_exposure/total,   "0.00%",  "right", bg)
        _dc(ws, rn, 7, row.notes,               None,     "left",  bg, False, "FF555555", 8, FONT_NAME, True)
        ws.row_dimensions[rn].height = 14


# ═══════════════════════════════════════════════════════════════════════════════
# Main builder
# ═══════════════════════════════════════════════════════════════════════════════

def build_report(output_path: str,
                 rollup_df: pd.DataFrame,
                 subclass_df: pd.DataFrame,
                 stock_df: pd.DataFrame,
                 sector_df: pd.DataFrame,
                 exposure_rows: list,
                 total_portfolio_value: float) -> str:
    """Build the full Excel report and save to output_path."""
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    _build_true_allocation(wb, rollup_df, subclass_df, total_portfolio_value)
    _build_subclass_sheet(wb, subclass_df, total_portfolio_value)
    _build_stock_rollup(wb, stock_df, total_portfolio_value)
    _build_sector_sheet(wb, sector_df, total_portfolio_value)
    _build_loothrough_sheet(wb, exposure_rows, total_portfolio_value)

    wb.save(output_path)
    print(f"[report_builder] ✅ Report saved: {output_path}")
    return output_path
